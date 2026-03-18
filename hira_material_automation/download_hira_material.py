from __future__ import annotations

import argparse
import csv
import json
import re
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from process_hira_mhtml_xls import (
    autofit_columns,
    ensure_directory,
    load_app_config,
    normalize_hira_export,
    resolve_config_paths,
    sha256sum,
)

try:
    KST = ZoneInfo("Asia/Seoul")
except Exception:
    KST = timezone(timedelta(hours=9))
NO_DATA_TEXT = "데이터가 없습니다"
BLOCKED_TEXT_PATTERNS = [
    "비정상 접속",
    "30분 후",
    "자동화된 접근",
]
RANGE_FILE_PATTERN = re.compile(
    r"^(?P<start>\d{6})_(?P<end>\d{6})__(?P<slug>.+)__(?P<code>\d+?)(?:__normalized)?$"
)
MASTER_HEADERS = [
    "기간",
    "연도",
    "중분류코드",
    "중분류명",
    "건강보험 청구량",
    "건강보험 청구금액",
    "의료급여 청구량",
    "의료급여 청구금액",
    "청구량 합계",
    "청구금액 합계",
]


@dataclass(frozen=True)
class Category:
    code: str
    name: str
    slug: str


@dataclass(frozen=True)
class MonthSpec:
    ym_dash: str
    ym_compact: str


@dataclass(frozen=True)
class PeriodSpec:
    start_month: MonthSpec
    end_month: MonthSpec

    @property
    def range_dash(self) -> str:
        return f"{self.start_month.ym_dash}~{self.end_month.ym_dash}"

    @property
    def range_compact(self) -> str:
        return f"{self.start_month.ym_compact}_{self.end_month.ym_compact}"


@dataclass(frozen=True)
class RangeFileInfo:
    path: Path
    start_month: str
    end_month: str
    category_code: str


def parse_year_month(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", value.strip())
    if not match:
        raise ValueError(f"Invalid month format: {value}. Expected YYYY-MM.")
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError(f"Invalid month value: {value}")
    return year, month


def format_year_month(year: int, month: int) -> MonthSpec:
    return MonthSpec(ym_dash=f"{year:04d}-{month:02d}", ym_compact=f"{year:04d}{month:02d}")


def compact_to_dash(value: str) -> str:
    if len(value) != 6 or not value.isdigit():
        raise ValueError(f"Invalid compact month format: {value}. Expected YYYYMM.")
    return f"{value[:4]}-{value[4:]}"


def shift_year_month(year: int, month: int, delta_months: int) -> tuple[int, int]:
    index = year * 12 + (month - 1) + delta_months
    shifted_year, shifted_month_index = divmod(index, 12)
    return shifted_year, shifted_month_index + 1


def iterate_months(start_month: str, end_month: str) -> list[MonthSpec]:
    start_year, start_value = parse_year_month(start_month)
    end_year, end_value = parse_year_month(end_month)
    current_year, current_month = start_year, start_value
    months: list[MonthSpec] = []
    while (current_year, current_month) <= (end_year, end_value):
        months.append(format_year_month(current_year, current_month))
        current_year, current_month = shift_year_month(current_year, current_month, 1)
    return months


def month_spec_key(month: MonthSpec) -> tuple[int, int]:
    return parse_year_month(month.ym_dash)


def today_kst() -> datetime:
    return datetime.now(KST)


def rolling_months(config: dict) -> list[MonthSpec]:
    window = int(config["sync"].get("rolling_window_months", 2))
    now = today_kst()
    end_year, end_month = now.year, now.month
    start_year, start_month = shift_year_month(end_year, end_month, -window)
    return iterate_months(f"{start_year:04d}-{start_month:02d}", f"{end_year:04d}-{end_month:02d}")


def make_period(start_month: str, end_month: str) -> PeriodSpec:
    start_year, start_value = parse_year_month(start_month)
    end_year, end_value = parse_year_month(end_month)
    if (start_year, start_value) > (end_year, end_value):
        raise ValueError(f"Invalid range: {start_month} is later than {end_month}.")
    return PeriodSpec(
        start_month=format_year_month(start_year, start_value),
        end_month=format_year_month(end_year, end_value),
    )


def count_inclusive_months(start_month: str, end_month: str) -> int:
    start_year, start_value = parse_year_month(start_month)
    end_year, end_value = parse_year_month(end_month)
    return (end_year - start_year) * 12 + (end_value - start_value) + 1


def split_into_periods(start_month: str, end_month: str, max_months_per_query: int) -> list[PeriodSpec]:
    if max_months_per_query < 1:
        raise ValueError("max_months_per_query must be at least 1.")
    periods: list[PeriodSpec] = []
    current_start = start_month
    while True:
        remaining_months = count_inclusive_months(current_start, end_month)
        if remaining_months <= max_months_per_query:
            periods.append(make_period(current_start, end_month))
            return periods
        start_year, start_value = parse_year_month(current_start)
        chunk_end_year, chunk_end_month = shift_year_month(start_year, start_value, max_months_per_query - 1)
        chunk_end = f"{chunk_end_year:04d}-{chunk_end_month:02d}"
        periods.append(make_period(current_start, chunk_end))
        next_year, next_month = shift_year_month(chunk_end_year, chunk_end_month, 1)
        current_start = f"{next_year:04d}-{next_month:02d}"


def parse_range_file_info(path: Path) -> RangeFileInfo | None:
    match = RANGE_FILE_PATTERN.fullmatch(path.stem)
    if not match:
        return None
    return RangeFileInfo(
        path=path,
        start_month=compact_to_dash(match.group("start")),
        end_month=compact_to_dash(match.group("end")),
        category_code=match.group("code"),
    )


def periods_overlap(start_month: str, end_month: str, other_start_month: str, other_end_month: str) -> bool:
    start_key = parse_year_month(start_month)
    end_key = parse_year_month(end_month)
    other_start_key = parse_year_month(other_start_month)
    other_end_key = parse_year_month(other_end_month)
    return not (end_key < other_start_key or other_end_key < start_key)


def purge_overlapping_outputs(
    raw_directory: Path,
    monthly_output_directory: Path,
    categories: list[Category],
    periods: list[PeriodSpec],
) -> dict[str, list[str]]:
    category_codes = {category.code for category in categories}
    deleted_raw_files: list[str] = []
    deleted_monthly_files: list[str] = []

    def purge_directory(directory: Path, bucket: list[str]) -> None:
        for path in directory.iterdir():
            if not path.is_file() or path.name == ".gitkeep":
                continue
            info = parse_range_file_info(path)
            if info is None or info.category_code not in category_codes:
                continue
            if any(
                periods_overlap(
                    info.start_month,
                    info.end_month,
                    period.start_month.ym_dash,
                    period.end_month.ym_dash,
                )
                for period in periods
            ):
                path.unlink()
                bucket.append(str(path))

    purge_directory(raw_directory, deleted_raw_files)
    purge_directory(monthly_output_directory, deleted_monthly_files)
    return {
        "deleted_raw_files": deleted_raw_files,
        "deleted_monthly_files": deleted_monthly_files,
    }


def load_categories(config: dict) -> list[Category]:
    categories: list[Category] = []
    seen_codes: set[str] = set()
    for item in config.get("categories", []):
        code = str(item["code"]).strip()
        if code in seen_codes:
            continue
        seen_codes.add(code)
        categories.append(
            Category(
                code=code,
                name=str(item["name"]).strip(),
                slug=str(item.get("slug") or code).strip(),
            )
        )
    return categories


def filter_categories(categories: list[Category], selected_codes: list[str]) -> list[Category]:
    if not selected_codes:
        return categories
    selected = {code.strip() for code in selected_codes}
    return [category for category in categories if category.code in selected]


def browser_launch(playwright, browser_name: str, headless: bool):
    if browser_name == "chrome":
        return playwright.chromium.launch(channel="chrome", headless=headless)
    return playwright.chromium.launch(headless=headless)


def extract_table_preview(page, max_rows: int = 10) -> list[list[str]]:
    rows: list[list[str]] = []
    row_locator = page.locator("div.tblType02.data table tr")
    row_count = min(row_locator.count(), max_rows)
    for index in range(row_count):
        cells = [cell.inner_text().strip() for cell in row_locator.nth(index).locator("th, td").all()]
        if cells:
            rows.append(cells)
    return rows


def preview_has_no_data(rows: list[list[str]]) -> bool:
    if not rows:
        return True
    flattened = " ".join(cell for row in rows for cell in row)
    if NO_DATA_TEXT in flattened:
        return True
    return len(rows) < 3


def page_has_block_notice(page) -> bool:
    try:
        body_text = page.locator("body").inner_text(timeout=5000)
    except Exception:  # noqa: BLE001
        return False
    return any(pattern in body_text for pattern in BLOCKED_TEXT_PATTERNS)


def reset_page(page) -> None:
    try:
        page.goto("about:blank", wait_until="load", timeout=10000)
    except Exception:  # noqa: BLE001
        pass


def query_page(page, category: Category, period: PeriodSpec, config: dict) -> list[list[str]]:
    timeout_ms = int(config["download"].get("timeout_ms", 120000))
    page.goto(config["download"]["base_url"], wait_until="networkidle", timeout=timeout_ms)
    page.wait_for_selector("#mcatMdivCd", state="attached", timeout=timeout_ms)
    page.wait_for_selector("#mcatMdivCdNm", state="attached", timeout=timeout_ms)
    page.wait_for_selector("#searchWrd", state="attached", timeout=timeout_ms)
    page.wait_for_selector("#sYm", state="attached", timeout=timeout_ms)
    page.wait_for_selector("#eYm", state="attached", timeout=timeout_ms)
    page.wait_for_selector("#searchBtn", state="attached", timeout=timeout_ms)
    page.evaluate(
        """({code, name}) => {
            document.querySelector('#mcatMdivCd').value = code;
            document.querySelector('#mcatMdivCdNm').value = name;
            document.querySelector('#searchWrd').value = name;
            const searchDate = document.querySelector('#searchDate');
            if (searchDate) {
              searchDate.style.display = 'block';
            }
        }""",
        {"code": category.code, "name": category.name},
    )
    page.fill("#sYm", period.start_month.ym_dash)
    page.fill("#eYm", period.end_month.ym_dash)
    page.locator("#searchBtn").click()
    page.wait_for_load_state("networkidle", timeout=timeout_ms)
    page.wait_for_timeout(1000)
    return extract_table_preview(page)


def download_to_temporary_file(page, target_path: Path, timeout_ms: int) -> Path:
    temp_path = target_path.with_name(f"tmp__{target_path.name}")
    if temp_path.exists():
        temp_path.unlink()
    with page.expect_download(timeout=timeout_ms) as download_info:
        page.locator("#exlBtn").click()
    download = download_info.value
    download.save_as(str(temp_path))
    return temp_path


def replace_raw_file(temp_path: Path, target_path: Path) -> dict:
    new_hash = sha256sum(temp_path)
    previous_hash = sha256sum(target_path) if target_path.exists() else ""
    if previous_hash and previous_hash == new_hash:
        temp_path.unlink()
        return {
            "raw_changed": False,
            "raw_hash": previous_hash,
            "previous_raw_hash": previous_hash,
            "raw_file": str(target_path),
        }

    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path.replace(target_path)
    return {
        "raw_changed": True,
        "raw_hash": new_hash,
        "previous_raw_hash": previous_hash,
        "raw_file": str(target_path),
    }


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        return [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        return reader.fieldnames or [], rows


def normalize_month_file(source_path: Path, monthly_output_directory: Path, processing: dict) -> dict:
    return normalize_hira_export(
        source_path=source_path,
        output_directory=monthly_output_directory,
        required_keywords=processing["required_header_keywords"],
        minimum_keyword_matches=int(processing["minimum_keyword_matches"]),
        write_csv_file=bool(processing["write_csv"]),
        write_xlsx_file=bool(processing["write_xlsx"]),
    )


def is_no_data_normalization_error(exc: Exception) -> bool:
    message = str(exc)
    patterns = [
        "No data rows were found under the detected header block.",
        "Not enough rows to identify a HIRA table.",
        "No parseable HTML table was found.",
    ]
    return any(pattern in message for pattern in patterns)


def detect_output_action(previous_rows: list[dict[str, str]], current_rows: list[dict[str, str]]) -> str:
    if not previous_rows and current_rows:
        return "new"
    if previous_rows == current_rows:
        return "unchanged"
    return "updated"


def find_column(headers: list[str], includes: list[str], excludes: list[str] | None = None) -> str:
    excludes = excludes or []
    for header in headers:
        if all(token in header for token in includes) and all(token not in header for token in excludes):
            return header
    return ""


def parse_numeric(value: str) -> int:
    text = str(value or "").strip()
    if text in {"", "-"}:
        return 0
    text = text.replace(",", "")
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return int(float(text))
    return 0


def parse_year_from_period(value: str) -> int:
    match = re.search(r"(\d{4})", value)
    return int(match.group(1)) if match else 0


def parse_period_sort_key(value: str) -> tuple[int, int]:
    match = re.search(r"(\d{4})\D+(\d{1,2})", value)
    if not match:
        return (0, 0)
    return int(match.group(1)), int(match.group(2))


def build_header_map(headers: list[str]) -> dict[str, str]:
    return {
        "period": find_column(headers, ["기간"]),
        "code": find_column(headers, ["중분류코드"], ["분류명"]),
        "name": find_column(headers, ["분류명"]) or find_column(headers, ["중분류명"]),
        "health_qty": find_column(headers, ["건강보험", "청구량"]),
        "health_amount": find_column(headers, ["건강보험", "청구금액"]),
        "medical_qty": find_column(headers, ["의료급여", "청구량"]),
        "medical_amount": find_column(headers, ["의료급여", "청구금액"]),
    }


def transform_row(source_row: dict[str, str], header_map: dict[str, str]) -> dict[str, object]:
    period = source_row.get(header_map["period"], "")
    code = source_row.get(header_map["code"], "")
    name = source_row.get(header_map["name"], "")
    health_qty = parse_numeric(source_row.get(header_map["health_qty"], "0"))
    health_amount = parse_numeric(source_row.get(header_map["health_amount"], "0"))
    medical_qty = parse_numeric(source_row.get(header_map["medical_qty"], "0"))
    medical_amount = parse_numeric(source_row.get(header_map["medical_amount"], "0"))
    return {
        "기간": period,
        "연도": parse_year_from_period(period),
        "중분류코드": code,
        "중분류명": name,
        "건강보험 청구량": health_qty,
        "건강보험 청구금액": health_amount,
        "의료급여 청구량": medical_qty,
        "의료급여 청구금액": medical_amount,
        "청구량 합계": health_qty + medical_qty,
        "청구금액 합계": health_amount + medical_amount,
    }


def normalize_master_row(row: dict[str, object]) -> list[object]:
    return [row.get(header, "") for header in MASTER_HEADERS]


def load_category_master_rows(monthly_output_directory: Path, category: Category) -> tuple[list[dict[str, object]], int]:
    monthly_files = sorted(monthly_output_directory.glob(f"*__{category.slug}__{category.code}__normalized.csv"))
    keyed_rows: dict[tuple[str, str], dict[str, object]] = {}
    input_row_count = 0

    for csv_path in monthly_files:
        headers, rows = read_csv_rows(csv_path)
        if not headers:
            continue
        header_map = build_header_map(headers)
        if not all(header_map.values()):
            continue
        for row in rows:
            input_row_count += 1
            transformed = transform_row(row, header_map)
            key = (str(transformed["기간"]), str(transformed["중분류코드"]))
            keyed_rows[key] = transformed

    transformed_rows = sorted(
        keyed_rows.values(),
        key=lambda row: (parse_period_sort_key(str(row["기간"])), str(row["중분류코드"])),
    )
    duplicates_removed = max(input_row_count - len(transformed_rows), 0)
    return transformed_rows, duplicates_removed


def make_sheet_name(base_name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", base_name).strip() or "sheet"
    candidate = cleaned[:31]
    counter = 1
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def write_rows_to_sheet(sheet, rows: list[dict[str, object]]) -> None:
    sheet.append(MASTER_HEADERS)
    for row in rows:
        sheet.append(normalize_master_row(row))

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    numeric_headers = {
        "연도",
        "건강보험 청구량",
        "건강보험 청구금액",
        "의료급여 청구량",
        "의료급여 청구금액",
        "청구량 합계",
        "청구금액 합계",
    }
    for column_index, header in enumerate(MASTER_HEADERS, start=1):
        if header in numeric_headers:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = "#,##0"

    autofit_columns(sheet)


def clear_master_output_directory(master_output_directory: Path) -> None:
    ensure_directory(master_output_directory)
    for path in master_output_directory.iterdir():
        if path.name == ".gitkeep":
            continue
        if path.is_file():
            path.unlink()


def create_summary_workbook(
    master_output_directory: Path,
    category_sheets: list[tuple[str, list[dict[str, object]]]],
    all_rows: list[dict[str, object]],
) -> str:
    workbook_path = master_output_directory / "hira_material_summary.xlsx"
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "통합"
    write_rows_to_sheet(summary_sheet, all_rows)

    for sheet_name, rows in category_sheets:
        sheet = workbook.create_sheet(title=sheet_name)
        write_rows_to_sheet(sheet, rows)

    workbook.save(workbook_path)
    return str(workbook_path)


def rebuild_master_datasets(monthly_output_directory: Path, master_output_directory: Path, categories: list[Category]) -> dict:
    clear_master_output_directory(master_output_directory)
    category_reports: list[dict] = []
    category_sheets: list[tuple[str, list[dict[str, object]]]] = []
    all_rows: list[dict[str, object]] = []
    total_duplicates_removed = 0
    used_sheet_names = {"통합"}

    for category in categories:
        rows, duplicates_removed = load_category_master_rows(monthly_output_directory, category)
        total_duplicates_removed += duplicates_removed
        if not rows:
            continue
        sheet_name = make_sheet_name(f"{category.code}_{category.name}", used_sheet_names)
        category_reports.append(
            {
                "category_code": category.code,
                "category_name": category.name,
                "sheet_name": sheet_name,
                "duplicates_removed": duplicates_removed,
                "row_count": len(rows),
            }
        )
        category_sheets.append((sheet_name, rows))
        all_rows.extend(rows)

    keyed_all_rows: dict[tuple[str, str], dict[str, object]] = {}
    for row in all_rows:
        key = (str(row["기간"]), str(row["중분류코드"]))
        keyed_all_rows[key] = row

    combined_rows = sorted(
        keyed_all_rows.values(),
        key=lambda row: (str(row["중분류코드"]), parse_period_sort_key(str(row["기간"]))),
    )

    combined_report = {}
    summary_workbook_path = ""
    if combined_rows:
        combined_report = {"row_count": len(combined_rows)}
        summary_workbook_path = create_summary_workbook(master_output_directory, category_sheets, combined_rows)

    return {
        "generated_at": today_kst().isoformat(timespec="seconds"),
        "category_reports": category_reports,
        "all_categories": combined_report,
        "duplicates_removed": total_duplicates_removed,
        "summary_workbook_path": summary_workbook_path,
    }


def run_category_period(
    page,
    category: Category,
    period: PeriodSpec,
    config: dict,
    raw_directory: Path,
    monthly_output_directory: Path,
    refresh_existing: bool,
    force: bool,
) -> dict:
    processing = config["processing"]
    timeout_ms = int(config["download"].get("timeout_ms", 120000))
    max_attempts = max(1, int(config["download"].get("max_attempts", 3)))
    retry_backoff_seconds = float(config["download"].get("retry_backoff_seconds", 4.0))
    file_name = f"{period.range_compact}__{category.slug}__{category.code}.xlsx"
    target_raw_path = raw_directory / file_name
    previous_csv_path = monthly_output_directory / f"{target_raw_path.stem}__normalized.csv"
    _, previous_rows = read_csv_rows(previous_csv_path)

    result = {
        "category_code": category.code,
        "category_name": category.name,
        "target_range": period.range_dash,
        "raw_file": str(target_raw_path),
        "monthly_csv": str(previous_csv_path),
        "action": "skipped",
        "table_preview": [],
        "no_data": False,
        "attempts": 0,
    }

    should_download = force or refresh_existing or not target_raw_path.exists()
    if not should_download:
        if target_raw_path.exists() and not previous_csv_path.exists():
            try:
                result["normalized"] = normalize_month_file(target_raw_path, monthly_output_directory, processing)
                _, current_rows = read_csv_rows(previous_csv_path)
                result["action"] = detect_output_action(previous_rows, current_rows)
                result["monthly_xlsx"] = str(monthly_output_directory / f"{target_raw_path.stem}__normalized.xlsx")
            except ValueError as exc:
                if is_no_data_normalization_error(exc):
                    result["no_data"] = True
                    result["action"] = "no_data_existing_raw"
                else:
                    raise
        else:
            result["action"] = "skipped_existing"
        return result

    for attempt in range(1, max_attempts + 1):
        result["attempts"] = attempt
        try:
            preview_rows = query_page(page, category, period, config)
            result["table_preview"] = preview_rows

            if page_has_block_notice(page):
                result["action"] = "blocked_keep_existing" if target_raw_path.exists() else "blocked"
                result["warning"] = "HIRA access block notice was detected."
                if attempt < max_attempts:
                    reset_page(page)
                    time.sleep(retry_backoff_seconds * attempt)
                    continue
                return result

            if preview_has_no_data(preview_rows):
                result["no_data"] = True
                result["action"] = "no_data_keep_existing" if target_raw_path.exists() else "no_data"
                return result

            try:
                temp_path = download_to_temporary_file(page, target_raw_path, timeout_ms)
            except PlaywrightTimeoutError:
                result["action"] = "download_timeout_keep_existing" if target_raw_path.exists() else "download_timeout"
                result["warning"] = "The export download timed out."
                if attempt < max_attempts:
                    reset_page(page)
                    time.sleep(retry_backoff_seconds * attempt)
                    continue
                return result
            break
        except PlaywrightTimeoutError:
            result["action"] = "query_timeout_keep_existing" if target_raw_path.exists() else "query_timeout"
            result["warning"] = "The HIRA page query timed out."
            if attempt < max_attempts:
                reset_page(page)
                time.sleep(retry_backoff_seconds * attempt)
                continue
            return result
    else:
        return result

    raw_status = replace_raw_file(temp_path, target_raw_path)
    result.update(raw_status)

    current_csv_path = monthly_output_directory / f"{target_raw_path.stem}__normalized.csv"
    need_normalize = force or raw_status["raw_changed"] or not current_csv_path.exists()
    if need_normalize:
        try:
            result["normalized"] = normalize_month_file(target_raw_path, monthly_output_directory, processing)
        except ValueError as exc:
            if is_no_data_normalization_error(exc):
                result["no_data"] = True
                result["action"] = "no_data"
                return result
            raise

    _, current_rows = read_csv_rows(current_csv_path)
    result["action"] = detect_output_action(previous_rows, current_rows)
    result["monthly_csv"] = str(current_csv_path)
    result["monthly_xlsx"] = str(monthly_output_directory / f"{target_raw_path.stem}__normalized.xlsx")
    return result


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Download HIRA treatment-material claim stats and normalize them.")
    parser.add_argument("--config", type=Path, default=None, help="Path to config.json")
    parser.add_argument("--browser", choices=["chromium", "chrome"], default="chromium", help="Browser runtime")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    parser.add_argument("--headed", action="store_true", help="Run browser with a visible window")
    parser.add_argument("--force", action="store_true", help="Re-download even if files already exist")
    parser.add_argument("--skip-process", action="store_true", help="Only download raw xlsx files")
    parser.add_argument("--mode", choices=["rolling", "backfill", "range"], default=None, help="Sync mode")
    parser.add_argument("--start-month", type=str, default="", help="Start month in YYYY-MM format")
    parser.add_argument("--end-month", type=str, default="", help="End month in YYYY-MM format")
    parser.add_argument("--category-code", action="append", default=None, help="Optional category code filter")
    return parser


def select_periods(config: dict, mode: str, start_month: str, end_month: str) -> list[PeriodSpec]:
    max_months_per_query = int(config["sync"].get("max_months_per_query", 36))

    if start_month and end_month:
        return split_into_periods(start_month, end_month, max_months_per_query)

    if mode == "backfill":
        return split_into_periods(
            config["sync"]["backfill_start_month"],
            config["sync"]["backfill_end_month"],
            max_months_per_query,
        )

    if mode == "range":
        raise ValueError("Range mode requires both --start-month and --end-month.")

    now = today_kst()
    end_year, end_month_value = now.year, now.month
    rolling_strategy = str(config["sync"].get("rolling_strategy", "window")).strip().lower()
    if rolling_strategy == "current_year_replace":
        rolling_start = f"{end_year:04d}-01"
    else:
        rolling_window = int(config["sync"].get("rolling_window_months", 2))
        start_year, start_month_value = shift_year_month(end_year, end_month_value, -rolling_window)
        rolling_start = f"{start_year:04d}-{start_month_value:02d}"
    rolling_end = f"{end_year:04d}-{end_month_value:02d}"
    rolling_min_month = str(config["sync"].get("rolling_min_month", "")).strip()
    if rolling_min_month and parse_year_month(rolling_end) < parse_year_month(rolling_min_month):
        return []
    if rolling_min_month and parse_year_month(rolling_start) < parse_year_month(rolling_min_month):
        rolling_start = rolling_min_month
    return [make_period(rolling_start, rolling_end)]


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    config, config_root = load_app_config(args.config)
    config = resolve_config_paths(config, config_root)

    mode = args.mode or config["sync"].get("default_mode", "rolling")
    periods = select_periods(config, mode, args.start_month.strip(), args.end_month.strip())
    categories = filter_categories(load_categories(config), args.category_code or [])
    if not periods:
        print("No target periods were selected.")
        return 0
    if not categories:
        print("No categories were selected.")
        return 1

    paths = config["paths"]
    raw_directory = ensure_directory(paths["raw_directory"])
    output_directory = ensure_directory(paths["output_directory"])
    monthly_output_directory = ensure_directory(output_directory / "monthly")
    master_output_directory = ensure_directory(output_directory / "master")
    log_directory = ensure_directory(paths["log_directory"])
    run_log_path = log_directory / "last_download_run.json"
    master_report_path = log_directory / "latest_master_report.json"

    if args.headed:
        headless = False
    elif args.headless:
        headless = True
    else:
        headless = bool(config["download"].get("headless", True))

    refresh_existing = mode == "rolling"
    delay_seconds = float(config["download"].get("request_delay_seconds", 0))
    results: list[dict] = []
    failures: list[dict] = []
    deleted_outputs = {"deleted_raw_files": [], "deleted_monthly_files": []}

    if mode == "rolling" and str(config["sync"].get("rolling_strategy", "window")).strip().lower() == "current_year_replace":
        deleted_outputs = purge_overlapping_outputs(raw_directory, monthly_output_directory, categories, periods)

    with sync_playwright() as playwright:
        browser = browser_launch(playwright, args.browser, headless)
        context = browser.new_context(locale="ko-KR", accept_downloads=True)

        for period in periods:
            for category in categories:
                page = context.new_page()
                try:
                    item = run_category_period(
                        page=page,
                        category=category,
                        period=period,
                        config=config,
                        raw_directory=raw_directory,
                        monthly_output_directory=monthly_output_directory,
                        refresh_existing=refresh_existing,
                        force=args.force,
                    )
                    if args.skip_process:
                        item.pop("normalized", None)
                    results.append(item)
                except Exception as exc:  # noqa: BLE001
                    failures.append(
                        {
                            "category_code": category.code,
                            "category_name": category.name,
                            "target_range": period.range_dash,
                            "error": str(exc),
                        }
                    )
                finally:
                    page.close()
                if delay_seconds > 0:
                    time.sleep(delay_seconds)

        context.close()
        browser.close()

    master_report = rebuild_master_datasets(monthly_output_directory, master_output_directory, categories)
    run_summary = {
        "mode": mode,
        "periods": [period.range_dash for period in periods],
        "processed": len(results),
        "failure_count": len(failures),
        "failures": failures,
        "deleted_outputs": deleted_outputs,
        "master_report": master_report,
        "results": results,
    }
    run_log_path.write_text(json.dumps(run_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    master_report_path.write_text(json.dumps(master_report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(run_summary, ensure_ascii=False, indent=2)
    )
    return 1 if failures and not results else 0


if __name__ == "__main__":
    raise SystemExit(main())



