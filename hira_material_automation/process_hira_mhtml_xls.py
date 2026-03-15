from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CONFIG = {
    "paths": {
        "inbox_directory": "./inbox",
        "raw_directory": "./raw",
        "archive_directory": "./archive",
        "output_directory": "./output",
        "log_directory": "./logs",
    },
    "processing": {
        "file_extensions": [".xlsx", ".xls", ".html", ".mhtml"],
        "required_header_keywords": ["\uc911\ubd84\ub958", "\uccad\uad6c"],
        "minimum_keyword_matches": 2,
        "archive_inbox_files": True,
        "write_csv": True,
        "write_xlsx": True,
    },
}

TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1")


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[str]]
    matched_keywords: int
    source_kind: str


def load_app_config(config_path: Path | None = None) -> tuple[dict, Path]:
    config_path = config_path or Path(__file__).with_name("config.json")
    loaded = {}
    if config_path.exists():
        loaded = json.loads(config_path.read_text(encoding="utf-8-sig"))

    merged = json.loads(json.dumps(DEFAULT_CONFIG))
    for section, values in loaded.items():
        if isinstance(values, dict) and isinstance(merged.get(section), dict):
            merged[section].update(values)
        else:
            merged[section] = values
    return merged, config_path.resolve().parent


def resolve_config_paths(config: dict, config_root: Path) -> dict:
    resolved = json.loads(json.dumps(config))
    for key, value in resolved["paths"].items():
        path = Path(value)
        resolved["paths"][key] = (config_root / path).resolve() if not path.is_absolute() else path.resolve()
    return resolved


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_number_like(value: str) -> bool:
    candidate = value.replace(",", "").strip()
    if not candidate:
        return False
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", candidate))


def is_header_like(row: list[str]) -> bool:
    filled = [cell for cell in row if cell]
    if not filled:
        return False
    if len(filled) == 1 and len(filled[0]) > 12:
        return False
    string_like = sum(0 if is_number_like(cell) else 1 for cell in filled)
    return string_like / len(filled) >= 0.6


def trim_grid(grid: list[list[str]]) -> list[list[str]]:
    rows = [[normalize_text(cell) for cell in row] for row in grid]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        return []

    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    keep_cols = [idx for idx in range(width) if any(row[idx] for row in padded)]
    return [[row[idx] for idx in keep_cols] for row in padded]


def dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    final_headers: list[str] = []
    for index, header in enumerate(headers, start=1):
        base = header or f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        final_headers.append(base if count == 1 else f"{base}_{count}")
    return final_headers


def forward_fill_header_row(row: list[str]) -> list[str]:
    filled: list[str] = []
    last_value = ""
    for cell in row:
        value = normalize_text(cell)
        if value:
            last_value = value
            filled.append(value)
        else:
            filled.append(last_value)
    return filled


def combine_headers(header_rows: list[list[str]]) -> list[str]:
    width = max(len(row) for row in header_rows)
    prepared_rows = [forward_fill_header_row(row + [""] * (width - len(row))) for row in header_rows]
    combined: list[str] = []
    for col_idx in range(width):
        parts: list[str] = []
        for row in prepared_rows:
            value = normalize_text(row[col_idx])
            if value and value not in parts:
                parts.append(value)
        header = " | ".join(parts)
        header = header.replace("\ubcf4\ud5d8\uc790\uad6c\ubd84 | ", "")
        combined.append(normalize_text(header))
    return dedupe_headers(combined)


def read_text_file(path: Path) -> str:
    for encoding in TEXT_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def html_table_to_grid(table) -> list[list[str]]:
    grid: list[list[str | None]] = []
    row_index = 0

    for tr in table.find_all("tr"):
        while len(grid) <= row_index:
            grid.append([])

        col_index = 0
        while col_index < len(grid[row_index]) and grid[row_index][col_index] is not None:
            col_index += 1

        for cell in tr.find_all(["th", "td"]):
            while len(grid[row_index]) <= col_index:
                grid[row_index].append(None)
            while grid[row_index][col_index] is not None:
                col_index += 1
                while len(grid[row_index]) <= col_index:
                    grid[row_index].append(None)

            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            text = normalize_text(cell.get_text(" ", strip=True))

            for row_offset in range(rowspan):
                target_row = row_index + row_offset
                while len(grid) <= target_row:
                    grid.append([])
                while len(grid[target_row]) < col_index + colspan:
                    grid[target_row].append(None)
                for col_offset in range(colspan):
                    grid[target_row][col_index + col_offset] = text

            col_index += colspan
        row_index += 1

    return trim_grid([[cell or "" for cell in row] for row in grid])


def workbook_to_grid(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([normalize_text(cell) for cell in row])
    workbook.close()
    return trim_grid(rows)


def find_header_start(grid: list[list[str]], required_keywords: list[str]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(grid):
        row_text = " ".join(row)
        score = sum(keyword in row_text for keyword in required_keywords)
        if score > best_score:
            best_score = score
            best_index = idx
    start = best_index
    while start > 0 and is_header_like(grid[start - 1]) and sum(bool(cell) for cell in grid[start - 1]) >= 2:
        start -= 1
        if best_index - start >= 2:
            break
    return start


def build_parsed_table(grid: list[list[str]], required_keywords: list[str], source_kind: str) -> ParsedTable:
    if len(grid) < 2:
        raise ValueError("Not enough rows to identify a HIRA table.")

    header_start = find_header_start(grid, required_keywords)
    header_end = header_start
    while header_end + 1 < len(grid) and is_header_like(grid[header_end + 1]):
        header_end += 1
        if header_end - header_start >= 2:
            break

    header_rows = grid[header_start : header_end + 1]
    headers = combine_headers(header_rows)

    rows: list[list[str]] = []
    for row in grid[header_end + 1 :]:
        padded = row + [""] * (len(headers) - len(row))
        padded = padded[: len(headers)]
        if any(padded):
            rows.append(padded)

    if not rows:
        raise ValueError("No data rows were found under the detected header block.")

    header_text = " ".join(headers)
    matched_keywords = sum(keyword in header_text for keyword in required_keywords)
    return ParsedTable(headers=headers, rows=rows, matched_keywords=matched_keywords, source_kind=source_kind)


def parse_source_file(path: Path, required_keywords: list[str], minimum_keyword_matches: int) -> ParsedTable:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        parsed = build_parsed_table(workbook_to_grid(path), required_keywords, "xlsx")
    else:
        html = read_text_file(path)
        soup = BeautifulSoup(html, "html.parser")
        candidates: list[ParsedTable] = []
        for table in soup.find_all("table"):
            grid = html_table_to_grid(table)
            if len(grid) < 2:
                continue
            try:
                candidates.append(build_parsed_table(grid, required_keywords, "html"))
            except ValueError:
                continue
        if not candidates:
            raise ValueError("No parseable HTML table was found.")
        candidates.sort(key=lambda item: (item.matched_keywords, len(item.rows), len(item.headers)), reverse=True)
        parsed = candidates[0]

    if required_keywords and parsed.matched_keywords < minimum_keyword_matches:
        raise ValueError(
            f"Not enough HIRA header keywords matched. Found {parsed.matched_keywords}, required {minimum_keyword_matches}."
        )
    return parsed


def sha256sum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def autofit_columns(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(normalize_text(cell.value)) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 10), 40)


def write_normalized_xlsx(path: Path, headers: list[str], rows: list[list[str]], metadata: dict) -> None:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "data"
    data_sheet.append(headers)
    for row in rows:
        data_sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    autofit_columns(data_sheet)

    meta_sheet = workbook.create_sheet("meta")
    meta_sheet.append(["key", "value"])
    for key, value in metadata.items():
        meta_sheet.append([key, value])
    for cell in meta_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    autofit_columns(meta_sheet)

    workbook.save(path)


def normalize_hira_export(
    source_path: Path,
    output_directory: Path,
    required_keywords: list[str],
    minimum_keyword_matches: int,
    write_csv_file: bool = True,
    write_xlsx_file: bool = True,
) -> dict:
    parsed = parse_source_file(source_path, required_keywords, minimum_keyword_matches)
    stem = source_path.stem
    csv_path = output_directory / f"{stem}__normalized.csv"
    xlsx_path = output_directory / f"{stem}__normalized.xlsx"
    metadata = {
        "source_file": str(source_path),
        "processed_at": datetime.now().isoformat(timespec="seconds"),
        "sha256": sha256sum(source_path),
        "source_kind": parsed.source_kind,
        "row_count": len(parsed.rows),
        "column_count": len(parsed.headers),
        "matched_keywords": parsed.matched_keywords,
    }

    ensure_directory(output_directory)
    if write_csv_file:
        write_csv(csv_path, parsed.headers, parsed.rows)
    if write_xlsx_file:
        write_normalized_xlsx(xlsx_path, parsed.headers, parsed.rows, metadata)

    return {
        **metadata,
        "csv_path": str(csv_path) if write_csv_file else "",
        "xlsx_path": str(xlsx_path) if write_xlsx_file else "",
        "headers": parsed.headers,
        "preview_row": parsed.rows[0] if parsed.rows else [],
    }


def load_manifest(path: Path) -> dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {"processed": {}}


def save_manifest(path: Path, manifest: dict) -> None:
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def discover_files(sources: Iterable[Path], extensions: set[str]) -> list[Path]:
    discovered: list[Path] = []
    for source in sources:
        if source.is_file() and source.suffix.lower() in extensions:
            discovered.append(source.resolve())
        elif source.is_dir():
            for path in sorted(source.iterdir()):
                if path.is_file() and path.suffix.lower() in extensions:
                    discovered.append(path.resolve())
    return discovered


def archive_file(source_path: Path, archive_directory: Path) -> Path:
    ensure_directory(archive_directory)
    target = archive_directory / f"{datetime.now():%Y%m%d_%H%M%S}__{source_path.name}"
    shutil.move(str(source_path), str(target))
    return target


def process_file(
    source_path: Path,
    output_directory: Path,
    archive_directory: Path,
    manifest: dict,
    required_keywords: list[str],
    minimum_keyword_matches: int,
    archive_source: bool,
    write_csv_file: bool,
    write_xlsx_file: bool,
    force: bool,
) -> dict | None:
    file_hash = sha256sum(source_path)
    if not force and file_hash in manifest["processed"]:
        return None

    result = normalize_hira_export(
        source_path=source_path,
        output_directory=output_directory,
        required_keywords=required_keywords,
        minimum_keyword_matches=minimum_keyword_matches,
        write_csv_file=write_csv_file,
        write_xlsx_file=write_xlsx_file,
    )

    result["archived_to"] = ""
    if archive_source:
        result["archived_to"] = str(archive_file(source_path, archive_directory))

    manifest["processed"][file_hash] = result
    return result


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Normalize downloaded HIRA material exports into CSV/XLSX.")
    parser.add_argument("--config", type=Path, default=None, help="Path to config.json")
    parser.add_argument("--input", type=Path, nargs="*", default=None, help="Optional file or directory override")
    parser.add_argument("--force", action="store_true", help="Reprocess files even if already seen")
    return parser


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    config, config_root = load_app_config(args.config)
    config = resolve_config_paths(config, config_root)

    paths = config["paths"]
    processing = config["processing"]

    output_directory = ensure_directory(paths["output_directory"])
    archive_directory = ensure_directory(paths["archive_directory"])
    log_directory = ensure_directory(paths["log_directory"])
    manifest_path = log_directory / "processed_manifest.json"
    manifest = load_manifest(manifest_path)

    source_inputs = args.input or [paths["inbox_directory"]]
    files = discover_files(source_inputs, {ext.lower() for ext in processing["file_extensions"]})
    if not files:
        print("No candidate files found.")
        return 0

    failures: list[tuple[Path, str]] = []
    processed_count = 0

    for source_path in files:
        try:
            result = process_file(
                source_path=source_path,
                output_directory=output_directory,
                archive_directory=archive_directory,
                manifest=manifest,
                required_keywords=processing["required_header_keywords"],
                minimum_keyword_matches=int(processing["minimum_keyword_matches"]),
                archive_source=bool(processing["archive_inbox_files"]),
                write_csv_file=bool(processing["write_csv"]),
                write_xlsx_file=bool(processing["write_xlsx"]),
                force=args.force,
            )
            if result is None:
                print(f"SKIPPED {source_path} (already processed)")
            else:
                processed_count += 1
                print(f"PROCESSED {source_path}")
        except Exception as exc:  # noqa: BLE001
            failures.append((source_path, str(exc)))
            print(f"FAILED {source_path}: {exc}", file=sys.stderr)

    save_manifest(manifest_path, manifest)

    print(f"Processed: {processed_count}")
    print(f"Failed: {len(failures)}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())

